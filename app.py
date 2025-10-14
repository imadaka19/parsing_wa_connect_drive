# app.py
import os
import zipfile
import re
import io
import pickle
from datetime import datetime
from difflib import get_close_matches

import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request

# -----------------------
# CONFIG / SCOPES
# -----------------------
SCOPES = ['https://www.googleapis.com/auth/drive.file']
DEFAULT_OUTPUT_FILENAME = "Stock_Opname_Diskrepensi.xlsx"

st.set_page_config(page_title="Parser Stock Opname (WhatsApp) → Drive", layout="wide")

st.title("Parser Stock Opname (WhatsApp) → Google Drive")
st.markdown(
    """
    WAJIB UP FILE CREDENTIAL DAN TOKEN DI KIRI!!!!!\nFORMAT WA SEBELUM EXTRACT WAJIB BAHASA INGGRIS DAN FORMAT WAKTU 24H KHUSUS ANDRO
    Upload file ZIP hasil export chat WhatsApp, aplikasi akan:
    - Mengekstrak file chat (.txt) dan gambar
    - Mem-parse format template: LOC, BIN, PN, SN, QTY EMRO, QTY ACTUAL, REMARK(S)
    - Upload gambar ke Google Drive (nama file: LOC-BIN-PN-SN-TANGGAL)
    - Export hasil ke Excel dan siapkan link download
    """
)

# -----------------------
# Sidebar: credentials & settings
# -----------------------
st.sidebar.header("Pengaturan Google Drive / OAuth WAJIB UPLOAD")

col1, col2 = st.sidebar.columns(2)
with col1:
    uploaded_credentials = st.sidebar.file_uploader("Upload credentials.json (OAuth client) download di https://drive.google.com/file/d/11qM5KzOrPjsHd_ck46KWVYOhlz6jHM1c/view?usp=drive_link", type=["json"])
    
with col2:
    uploaded_token = st.sidebar.file_uploader("Upload token.pickle (opsional, hasil autentikasi lokal) download di https://drive.google.com/file/d/1B7Jx4bav9HsGqd-f5DKNvi_omNY9v7tq/view?usp=drive_link", type=["pickle", "dat", "pkl"])

# folder_id = st.sidebar.text_input("Folder ID Google Drive tujuan", value="", help="ID folder di Google Drive tempat gambar akan diupload")

st.sidebar.markdown("---")
st.sidebar.markdown("**Petunjuk singkat**:\n WAJIB UP FILE CREDENTIAL DAN TOKEN DI KIRI!!!!!\nFORMAT WA SEBELUM EXTRACT WAJIB BAHASA INGGRIS DAN FORMAT WAKTU 24H KHUSUS ANDRO\n\n1. Jika deploy di cloud, lakukan autentikasi **sekali** di lokal untuk menghasilkan `token.pickle` menggunakan `credentials.json`.\n2. Upload `credentials.json` dan `token.pickle` di sidebar sebelum klik Proses.\n\nJika tidak punya token.pickle, jalankan versi lokal untuk menghasilkan file tersebut.")
st.sidebar.markdown("---")

# -----------------------
# Helper functions
# -----------------------
def load_credentials_from_uploaded(uploaded_credentials, uploaded_token):
    """
    Mengembalikan creds object yang dapat dipakai oleh googleapiclient.
    Prefer: token.pickle (uploaded) -> token.pickle on disk -> run local flow (only if running locally)
    """
    creds = None

    # save credentials.json to disk if uploaded
    if uploaded_credentials is not None:
        creds_path = os.path.join(".", "credentials_uploaded.json")
        with open(creds_path, "wb") as f:
            f.write(uploaded_credentials.getbuffer())
        credentials_file = creds_path
    else:
        credentials_file = "credentials.json" if os.path.exists("credentials.json") else None

    # token : prefer uploaded_token (file uploader)
    if uploaded_token is not None:
        token_path = os.path.join(".", "token_uploaded.pickle")
        with open(token_path, "wb") as f:
            f.write(uploaded_token.getbuffer())
        token_file = token_path
    else:
        token_file = "token.pickle" if os.path.exists("token.pickle") else None

    # try load token
    if token_file and os.path.exists(token_file):
        with open(token_file, "rb") as token:
            creds = pickle.load(token)

    # refresh or run local flow if possible
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
            except Exception as e:
                st.warning(f"Gagal refresh token: {e}")
                creds = None
        else:
            # only permit running local flow if credentials_file exists AND we are running locally (not enforced here)
            if credentials_file:
                try:
                    flow = InstalledAppFlow.from_client_secrets_file(credentials_file, SCOPES)
                    # IMPORTANT: When deployed to Cloud, run_local_server won't work.
                    # We attempt to run console flow if run_local_server is not appropriate.
                    # But in general, user should generate token locally and upload token.pickle to the cloud.
                    creds = flow.run_local_server(port=0)
                    # save token for future runs
                    with open("token.pickle", "wb") as token:
                        pickle.dump(creds, token)
                except Exception as e:
                    st.error(
                        "Autentikasi otomatis gagal. Jika Anda menjalankan ini di Streamlit Cloud, "
                        "silakan autentikasi di mesin lokal untuk membuat token.pickle, lalu upload token.pickle di sidebar.\n\n"
                        f"Error detail: {e}"
                    )
                    raise
            else:
                st.error("credentials.json tidak ditemukan. Upload credentials.json di sidebar atau taruh file credentials.json di folder aplikasi.")
                raise FileNotFoundError("credentials.json tidak ditemukan.")
    return creds

def build_drive_service(creds):
    return build('drive', 'v3', credentials=creds, cache_discovery=False)

folder_id = '1xBFx4yA7jV5OvevS1nG3HKON5glx1YUa'

def upload_to_drive(service, folder_id, img_path, entry):
    try:
        loc = (entry.get("LOC", "") or "").replace("/", "-").strip()
        bin_ = (entry.get("BIN", "") or "").replace("/", "-").strip()
        pn = (entry.get("PN", "") or "").replace("/", "-").strip()
        sn = (entry.get("SN", "") or "").replace("/", "-").strip() or "NO_SN"
        tanggal = (entry.get("Tanggal", "") or "").split(" ")[0].replace("/", "-").strip()
        parts = [p for p in [loc, bin_, pn, sn, tanggal] if p]
        filename = "-".join(parts) if parts else os.path.basename(img_path)
        ext = os.path.splitext(img_path)[1]
        filename = f"{filename}{ext}"
        file_metadata = {"name": filename}
        if folder_id:
            file_metadata["parents"] = [folder_id]
        media = MediaFileUpload(img_path)
        uploaded = service.files().create(body=file_metadata, media_body=media, fields="id").execute()
        # set public reader
        service.permissions().create(fileId=uploaded["id"], body={"type": "anyone", "role": "reader"}).execute()
        return f"https://drive.google.com/uc?id={uploaded['id']}"
    except Exception as e:
        st.error(f"Upload error for {os.path.basename(img_path)}: {e}")
        return ""

def extract_zip_to_dir(zip_bytes_io, extract_dir):
    with zipfile.ZipFile(zip_bytes_io) as zip_ref:
        zip_ref.extractall(extract_dir)

def find_text_chat_file(extract_dir):
    for root, _, files in os.walk(extract_dir):
        for f in files:
            if f.lower().endswith(".txt"):
                return os.path.join(root, f)
    return None

def index_images(extract_dir):
    image_index = {}
    for root, _, files in os.walk(extract_dir):
        for f in files:
            if f.lower().endswith((".jpg", ".jpeg", ".png")):
                image_index[f.lower()] = os.path.join(root, f)
    return image_index

def find_image_fuzzy(name, image_index):
    if not name:
        return None
    name_low = name.lower()
    if name_low in image_index:
        return image_index[name_low]
    candidates = get_close_matches(name_low, image_index.keys(), n=1, cutoff=0.5)
    if candidates:
        return image_index[candidates[0]]
    return None

def parse_chat_file(chat_file_path, image_index):
    entries = []
    current = {}
    image_file = None
    current_sender = ""
    current_date = ""

    def save_current():
        if current.get("LOC"):
            entries.append({
                "Tanggal": current_date,
                "Pengirim": current_sender,
                "LOC": current.get("LOC",""),
                "BIN": current.get("BIN",""),
                "PN": current.get("PN",""),
                "SN": current.get("SN",""),
                "Qty EMRO": current.get("QTY EMRO",""),
                "Qty ACTUAL": current.get("QTY ACTUAL",""),
                "Remarkable": current.get("REMARKABLE",""),
                "PHOTO FILE": image_file or "",
                "PHOTO LINK": ""
            })

    pattern_msg = re.compile(r"^(\d{1,2}/\d{1,2}/\d{2,4}), (\d{1,2}:\d{2}) - ([^:]+):")

    with open(chat_file_path, encoding="utf-8", errors="ignore") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            m = pattern_msg.match(line)
            if m:
                save_current()
                current = {}
                current_date = f"{m.group(1)} {m.group(2)}"
                current_sender = m.group(3).strip()
                match_img = re.search(r'IMG-\d{8}-WA\d{4}.*?\.(?:jpg|jpeg|png)', line, re.IGNORECASE)
                image_file = match_img.group(0).strip() if match_img else None
                continue
            match_img = re.search(r'IMG-\d{8}-WA\d{4}.*?\.(?:jpg|jpeg|png)', line, re.IGNORECASE)
            if match_img:
                image_file = match_img.group(0).strip()
                continue
            if "<media omitted>" in line.lower():
                image_file = None
                continue
            if ":" in line:
                key, val = line.split(":", 1)
                key = key.strip().upper()
                val = val.strip()
                key_map = {
                    "QTY ACT": "QTY ACTUAL",
                    "QTY ACTUAL": "QTY ACTUAL",
                    "QTY EMRO": "QTY EMRO",
                    "REMARK": "REMARKABLE",
                    "REMARKS": "REMARKABLE",
                    "REMARK(S)": "REMARKABLE",
                }
                key = key_map.get(key, key)
                current[key] = val
    save_current()
    return entries

def create_excel_bytes(entries):
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Opname"
    headers = [
        "Tanggal", "Pengirim", "LOC", "BIN", "PN", "SN",
        "Qty EMRO", "Qty ACTUAL", "Remarkable",
        "PHOTO FILE", "PHOTO LINK"
    ]
    ws.append(headers)
    col_widths = {
        "A": 18, "B": 25, "C": 10, "D": 10, "E": 20, "F": 20,
        "G": 12, "H": 12, "I": 25, "J": 35, "K": 40
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    for idx, item in enumerate(entries, start=2):
        for col_idx, key in enumerate(headers[:-1], start=1):
            ws.cell(row=idx, column=col_idx, value=item.get(key, ""))
        ws.cell(row=idx, column=len(headers), value=item.get("PHOTO LINK",""))
    header_font = Font(bold=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

# -----------------------
# UI: Upload ZIP
# -----------------------
st.header("1) Upload file ZIP WhatsApp")
zip_file_u = st.file_uploader("Pilih file ZIP (hasil export chat + media)", type=["zip"], accept_multiple_files=False)

col_local = st.columns(2)
run_local_auth = st.checkbox("Jalankan autentikasi lokal jika token tidak ada (hanya untuk environment lokal)", value=False)

# -----------------------
# PROCESS BUTTON
# -----------------------
if st.button("Proses Data"):
    if not zip_file_u:
        st.warning("Upload file ZIP dulu.")
        st.stop()

    # prepare temp dirs
    tmp_dir = "tmp_extracted"
    if os.path.exists(tmp_dir):
        import shutil
        shutil.rmtree(tmp_dir)
    os.makedirs(tmp_dir, exist_ok=True)

    # write zip to BytesIO and extract
    zip_bytes = io.BytesIO(zip_file_u.read())
    try:
        extract_zip_to_dir(zip_bytes, tmp_dir)
    except Exception as e:
        st.error(f"Gagal ekstrak ZIP: {e}")
        st.stop()

    # index images + find chat file
    image_index = index_images(tmp_dir)
    chat_file = find_text_chat_file(tmp_dir)
    if not chat_file:
        st.error("File chat (.txt) tidak ditemukan di ZIP.")
        st.stop()

    st.success(f"File chat ditemukan: {chat_file}")
    st.info(f"Jumlah gambar terindeks: {len(image_index)}")

    # get creds
    try:
        creds = load_credentials_from_uploaded(uploaded_credentials, uploaded_token)
    except Exception:
        st.stop()

    # if cloud and run_local_auth is True -> warn
    if run_local_auth and ("RUNNING_LOCALLY" not in os.environ):
        st.warning("Opsi run_local_auth aktif — pastikan Anda menjalankan ini di mesin lokal. Jika di cloud, autentikasi lokal tidak akan berhasil.")
    service = None
    try:
        service = build_drive_service(creds)
    except Exception as e:
        st.error(f"Gagal membuat Drive service: {e}")
        st.stop()

    # parse chat
    entries = parse_chat_file(chat_file, image_index)
    st.write(f"🔎 Total entri ditemukan: {len(entries)}")

    # progress UI
    progress_bar = st.progress(0)
    status_text = st.empty()

    # upload images and fill PHOTO LINK
    for i, entry in enumerate(entries, start=1):
        photo_name = entry.get("PHOTO FILE","")
        if photo_name:
            img_path = find_image_fuzzy(photo_name, image_index)
            if img_path and os.path.exists(img_path):
                status_text.text(f"Uploading image {i}/{len(entries)} → {os.path.basename(img_path)}")
                link = upload_to_drive(service, folder_id, img_path, entry)
                entry["PHOTO LINK"] = link or "UPLOAD FAILED"
            else:
                entry["PHOTO LINK"] = "NOT FOUND"
        else:
            entry["PHOTO LINK"] = "-"
        progress_bar.progress(int(100 * i / max(len(entries), 1)))

    progress_bar.empty()
    status_text.success("Semua entri diproses.")

    # create excel bytes
    excel_bytes = create_excel_bytes(entries)

    # show table preview (first 20 rows)
    st.header("Preview hasil (maks 20 baris)")
    import pandas as pd
    preview_df = pd.DataFrame(entries)
    st.dataframe(preview_df.head(20))

    # download button
    st.download_button(
        label="Unduh Excel (.xlsx)",
        data=excel_bytes.getvalue(),
        file_name=DEFAULT_OUTPUT_FILENAME,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    st.success("Selesai ✅. Jika kamu deploy di Streamlit Cloud, pastikan credentials.json + token.pickle sudah ditambahkan ke repository (atau diupload melalui sidebar) sebelum memproses.")
